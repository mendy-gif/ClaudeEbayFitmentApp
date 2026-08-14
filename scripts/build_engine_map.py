#!/usr/bin/env python3
"""
Validate data/bmw_engine_map.json against the chassis reference (every (chassis,
trim) must have >=1 engine entry) and emit review formats:
  - data/bmw_engine_map.csv
  - data/bmw_engine_map.xlsx  (confidence-colored, if openpyxl present)

Run: python3 scripts/build_engine_map.py
"""
import csv
import json
import os
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REF = os.path.join(ROOT, "data", "bmw_chassis_reference.json")
EMAP = os.path.join(ROOT, "data", "bmw_engine_map.json")
CSV_OUT = os.path.join(ROOT, "data", "bmw_engine_map.csv")
XLSX_OUT = os.path.join(ROOT, "data", "bmw_engine_map.xlsx")
COLS = ["chassis_code", "trim", "engine_code", "engine", "fuel", "year_range", "confidence", "verified", "notes"]


def norm(s):
    return " ".join(str(s).lower().split())


def main():
    ref = json.load(open(REF, encoding="utf-8"))
    emap = json.load(open(EMAP, encoding="utf-8"))

    # Coverage: every (chassis, trim) in ref must appear in emap.
    emap_keys = {(norm(e["chassis_code"]), norm(e["trim"])) for e in emap}
    missing = []
    ref_pairs = 0
    for row in ref:
        for t in row["trims"]:
            ref_pairs += 1
            if (norm(row["chassis_code"]), norm(t)) not in emap_keys:
                missing.append(f"{row['chassis_code']} / {t}")

    # Engine-family tally (for Rule B grouping sanity).
    families = {}
    for e in emap:
        families[e["engine_code"]] = families.get(e["engine_code"], 0) + 1

    # Write CSV
    with open(CSV_OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLS)
        w.writeheader()
        for e in emap:
            row = dict(e)
            row["verified"] = "yes" if e.get("verified") else "no"
            w.writerow({c: row.get(c, "") for c in COLS})

    xlsx_ok = write_xlsx(emap)

    print(f"Engine-map entries: {len(emap)}")
    print(f"Reference (chassis,trim) pairs: {ref_pairs}")
    print(f"Distinct engine families: {len(families)}")
    conf = {}
    for e in emap:
        conf[e.get("confidence", "?")] = conf.get(e.get("confidence", "?"), 0) + 1
    print(f"Confidence: {conf}")
    top = sorted(families.items(), key=lambda kv: -kv[1])[:12]
    print("Top engine families: " + ", ".join(f"{k}({v})" for k, v in top))
    print(f"Wrote: data/bmw_engine_map.csv" + (", data/bmw_engine_map.xlsx" if xlsx_ok else ""))

    if missing:
        print(f"\nCOVERAGE GAPS ({len(missing)}) - (chassis,trim) in reference with NO engine entry:")
        for m in missing:
            print("  - " + m)
        sys.exit(1)
    print("Coverage: OK - every reference (chassis,trim) has an engine.")


def write_xlsx(emap):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False
    wb = Workbook(); ws = wb.active; ws.title = "BMW Engine Map"
    hfill = PatternFill("solid", fgColor="1F3864"); hfont = Font(bold=True, color="FFFFFF")
    cfill = {"high": PatternFill("solid", fgColor="C6EFCE"), "medium": PatternFill("solid", fgColor="FFEB9C"), "low": PatternFill("solid", fgColor="FFC7CE")}
    ws.append([c.replace("_", " ").title() for c in COLS])
    for c in range(1, len(COLS) + 1):
        ws.cell(row=1, column=c).fill = hfill; ws.cell(row=1, column=c).font = hfont
    for e in emap:
        row = dict(e); row["verified"] = "yes" if e.get("verified") else "no"
        ws.append([row.get(c, "") for c in COLS])
        conf = e.get("confidence", "")
        if conf in cfill:
            ws.cell(row=ws.max_row, column=COLS.index("confidence") + 1).fill = cfill[conf]
    widths = {"chassis_code": 14, "trim": 26, "engine_code": 14, "engine": 40, "fuel": 8, "year_range": 12, "confidence": 11, "verified": 9, "notes": 55}
    for i, c in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"
    wb.save(XLSX_OUT); return True


if __name__ == "__main__":
    main()
