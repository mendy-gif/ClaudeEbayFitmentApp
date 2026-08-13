#!/usr/bin/env python3
"""
Build derived formats from data/bmw_chassis_reference.json:
  - data/bmw_chassis_reference.csv  (source-of-truth, diffable)
  - data/bmw_chassis_reference.xlsx (review-friendly, if openpyxl is available)

Also validates the JSON (unique enough keys, year sanity) and prints a summary.
Run: python3 scripts/build_reference.py
"""
import csv
import json
import os
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSON_PATH = os.path.join(ROOT, "data", "bmw_chassis_reference.json")
CSV_PATH = os.path.join(ROOT, "data", "bmw_chassis_reference.csv")
XLSX_PATH = os.path.join(ROOT, "data", "bmw_chassis_reference.xlsx")

COLUMNS = ["chassis_code", "series", "body_style", "us_start_year", "us_end_year",
           "still_selling", "trims", "confidence", "verified", "notes", "sources"]


def load():
    with open(JSON_PATH, encoding="utf-8") as f:
        return json.load(f)


def validate(rows):
    problems = []
    for i, r in enumerate(rows):
        for col in ("chassis_code", "series", "us_start_year", "us_end_year", "trims"):
            if col not in r:
                problems.append(f"row {i} ({r.get('chassis_code','?')}): missing '{col}'")
        s, e = r.get("us_start_year"), r.get("us_end_year")
        if isinstance(s, int) and isinstance(e, int) and s > e:
            problems.append(f"row {i} ({r.get('chassis_code')}): start {s} > end {e}")
        if not r.get("trims"):
            problems.append(f"row {i} ({r.get('chassis_code')}): empty trims")
    return problems


def flatten(r):
    out = dict(r)
    out["trims"] = " | ".join(r.get("trims", []))
    out["sources"] = " | ".join(r.get("sources", []))
    out["still_selling"] = "yes" if r.get("still_selling") else "no"
    return {c: out.get(c, "") for c in COLUMNS}


def write_csv(rows):
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow(flatten(r))


def write_xlsx(rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed - skipping .xlsx (CSV still written).")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "BMW Chassis Reference"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    conf_fill = {
        "high": PatternFill("solid", fgColor="C6EFCE"),
        "medium": PatternFill("solid", fgColor="FFEB9C"),
        "low": PatternFill("solid", fgColor="FFC7CE"),
    }

    ws.append([c.replace("_", " ").title() for c in COLUMNS])
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for r in rows:
        fr = flatten(r)
        ws.append([fr[c] for c in COLUMNS])
        row_idx = ws.max_row
        conf = r.get("confidence", "")
        if conf in conf_fill:
            ws.cell(row=row_idx, column=COLUMNS.index("confidence") + 1).fill = conf_fill[conf]

    widths = {"chassis_code": 14, "series": 22, "body_style": 26, "us_start_year": 8,
              "us_end_year": 8, "still_selling": 8, "trims": 46, "confidence": 10,
              "verified": 8, "notes": 60, "sources": 40}
    for i, c in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    wb.save(XLSX_PATH)
    return True


def main():
    rows = load()
    problems = validate(rows)
    write_csv(rows)
    xlsx_ok = write_xlsx(rows)

    print(f"Rows: {len(rows)}")
    families = {}
    for r in rows:
        fam = r["series"].split()[0]
        families[fam] = families.get(fam, 0) + 1
    still = sum(1 for r in rows if r.get("still_selling"))
    conf_counts = {}
    for r in rows:
        conf_counts[r.get("confidence", "?")] = conf_counts.get(r.get("confidence", "?"), 0) + 1
    print(f"Still selling (2026): {still}")
    print(f"Confidence: {conf_counts}")
    print(f"Wrote: {os.path.relpath(CSV_PATH, ROOT)}" + (f", {os.path.relpath(XLSX_PATH, ROOT)}" if xlsx_ok else ""))
    if problems:
        print(f"\nVALIDATION PROBLEMS ({len(problems)}):")
        for p in problems:
            print("  - " + p)
        sys.exit(1)
    print("Validation: OK")


if __name__ == "__main__":
    main()
