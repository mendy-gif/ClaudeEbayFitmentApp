#!/usr/bin/env python3
"""Step 1 — build the part-number -> vehicle fitment table.

Reads an inventory workbook (Sold + Small Parts sheets by default), aggregates
every distinct vehicle each part number has been observed on, and writes
fitment_by_partnumber.csv.

Usage:
  python3 build_fitment_table.py --inventory /path/to/Inventory.xlsx \
      --out ../data/fitment_by_partnumber.csv
"""
import argparse
import csv
import os
from collections import defaultdict

import openpyxl

from fitment_common import part_keys, parse_car


def build(inventory_path, sheets, out_path):
    wb = openpyxl.load_workbook(inventory_path, read_only=True, data_only=True)
    part_vehicles = defaultdict(set)     # key7 -> {(year, make, model)}
    part_11forms = defaultdict(set)      # key7 -> {11-digit forms}
    observations = 0

    for name in sheets:
        if name not in wb.sheetnames:
            print(f"  (skipping missing sheet: {name!r})")
            continue
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        header = list(next(it))
        idx = {h: i for i, h in enumerate(header) if h is not None}
        ci = idx.get("Car")
        pcols = [idx[c] for c in ("Part #", "Part # (2)") if c in idx]
        for row in it:
            if all(c is None for c in row):
                continue
            veh = parse_car(row[ci]) if ci is not None else None
            keys, elevens = set(), set()
            for col in pcols:
                for k7, f11 in part_keys(row[col]):
                    keys.add(k7)
                    if f11:
                        elevens.add(f11)
            if keys and veh:
                observations += 1
                for k in keys:
                    part_vehicles[k].add(veh)
                    part_11forms[k].update(elevens)

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["part_number_7", "part_number_11_forms", "vehicle_count", "vehicles"])
        for k in sorted(part_vehicles, key=lambda k: -len(part_vehicles[k])):
            veh = sorted(part_vehicles[k])
            w.writerow([k, "|".join(sorted(part_11forms[k])), len(veh),
                        "; ".join(f"{y} {mk} {md}" for y, mk, md in veh)])

    print(f"parts: {len(part_vehicles)}   observations: {observations}   -> {out_path}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--inventory", required=True, help="path to inventory .xlsx")
    ap.add_argument("--sheets", nargs="+", default=["Small Parts", "Sold"])
    ap.add_argument("--out", default="../data/fitment_by_partnumber.csv")
    a = ap.parse_args()
    build(a.inventory, a.sheets, a.out)
